import sys
import pandas as pd
import os
import numpy as np
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
warnings.filterwarnings("ignore", category=FutureWarning, message="Setting an item of incompatible dtype is deprecated*")
from openpyxl import load_workbook
from openpyxl.styles import Font


class SHARED_FUNCTION:
	
	def __init__(self, path):
		self.PATH = path

	# 整合所有成本表
	def Concat_Cost_Sheet(self):

		DF_Cost_List = []

		for file in os.listdir(self.PATH):
			if "成本表" in file:
				file_path = os.path.join(self.PATH, file)

				Cost_Workbook = pd.ExcelFile(file_path)
				Cost_Sheet_Name = [sheet for sheet in Cost_Workbook.sheet_names if "成本表" in sheet]

				for sheet in Cost_Sheet_Name:
					DF_Cost_Sheet = pd.read_excel(file_path, sheet_name=sheet, header=3)
					DF_Cost_Sheet.columns = DF_Cost_Sheet.columns.str.replace("\n", "") 
					DF_Cost_List.append((DF_Cost_Sheet, file_path))

		if not DF_Cost_List: 
			raise ValueError(f"No valid '成本表' sheets found.")

		Summarized_Sheet = pd.concat([df for df, _ in DF_Cost_List], axis=0, sort=False)

		return Summarized_Sheet, DF_Cost_List

	# 考量重量、數量、長度後，計算MOQ。
	def MOQ_Counting(self, Quoting_Info):
		Quoting_Info["MOQ"] = 300 / Quoting_Info["Weight_Per_Pcs"]

		# Standardize 'Size' format and extract 'Length'
		Quoting_Info["Size"] = Quoting_Info["Size"].astype(str).str.replace("X", "x")
		Quoting_Info["Length"] = Quoting_Info["Size"].apply(lambda x: x.split("x")[1] if "x" in x else None)

		# Apply MOQ minimums based on Length
		Quoting_Info.loc[(Quoting_Info["Length"].astype(float) <= 100) & (Quoting_Info["MOQ"] < 100), "MOQ"] = 100 
		Quoting_Info.loc[(Quoting_Info["Length"].astype(float) > 100) & (Quoting_Info["MOQ"] < 50), "MOQ"] = 50

		# Round up MOQ to nearest 10
		Quoting_Info["MOQ"] = Quoting_Info["MOQ"].apply(lambda x: int(np.ceil(x / 10) * 10) if pd.notna(x) else np.nan)

		# Remove MOQ if it's less than or equal to Quantity
		Quoting_Info.loc[Quoting_Info["MOQ"] <= Quoting_Info["Quantity"], "MOQ"] = None

		# If MOQ > Quantity, update Quantity
		Quoting_Info.loc[Quoting_Info["MOQ"] > Quoting_Info["Quantity"], "Quantity"] = \
			Quoting_Info.loc[Quoting_Info["MOQ"] > Quoting_Info["Quantity"], "MOQ"]

		# Convert MOQ from M (thousands) to pcs
		Quoting_Info["MOQ"] = (Quoting_Info["MOQ"].fillna(0) * 1000).astype(int)

		return Quoting_Info

	#計算每份成本表的報價重量及支數，及本詢價單報價總重及總支數
	def Count_Quote_Weight(self):

		Summarized_Sheet, DF_Cost_List = self.Concat_Cost_Sheet()
		
		quote_total_weight = 0
		quote_item_amount = 0

		for df, file_path in DF_Cost_List:
			df = df.copy()
			df["孔穴牙長"] = df["孔穴牙長"].fillna("").astype(str)
			df = df[~df["孔穴牙長"].str.startswith("NQ", na=False)] #NQ的不計算


			# Drop rows with missing values in critical columns
			df.loc[:, "數量(M)"] = pd.to_numeric(df["數量(M)"], errors="coerce")
			df.loc[:, "加工價(單重/M)"] = pd.to_numeric(df["加工價(單重/M)"], errors="coerce")
			df.dropna(subset=["數量(M)", "加工價(單重/M)"], inplace=True)

			# Compute weight
			total_weight = (df["加工價(單重/M)"] * df["數量(M)"]).sum()

			quote_total_weight = quote_total_weight + total_weight
			quote_item_amount = quote_item_amount + df.shape[0]

			#顯示每一大類總重及支數
			print(f"\nFile Name: {os.path.basename(file_path)}\n-Total Weight: {total_weight:.2f} kg\n-Item Amount: {df.shape[0]}")
		
		#顯示此份詢價單總重
		print(f"\n報價總重: {quote_total_weight:.2f} kg")
		print(f"報價總支數: {quote_item_amount}支")

		return quote_total_weight, quote_item_amount
		

class FILL_RFQ_C006(SHARED_FUNCTION):

	def __init__(self, path):
		self.PATH = path
		super().__init__(path)

	def Load_Cost_Info(self):

		#取得所有成本表資料
		Summarized_Sheet, DF_Cost_List = self.Concat_Cost_Sheet()

		#選取需要欄位並重新命名
		column_map = {
			Summarized_Sheet.columns[0]: "NO.",
			"客戶代號": "Item_Code",
			"尺寸": "Size",
			"孔穴牙長": "Status",
			"MOQ": "MOQ",
			"數量(M)": "Quantity",
			"加工價(單重/M)": "Weight_Per_Pcs",
			"總計/M": "Price/M"
		}

		Summarized_Sheet = Summarized_Sheet[list(column_map.keys())].rename(columns=column_map)
		
		#將 MOQ 數量 單重 價錢 轉換成可運算之數字
		Summarized_Sheet.iloc[:, 4:] = Summarized_Sheet.iloc[:, 4:].apply(pd.to_numeric, errors="coerce")

		#移除空格行及品名行，僅保留客戶產品代號的報價資訊
		Quoting_Info = Summarized_Sheet[Summarized_Sheet["Item_Code"].astype(str).str.match(r'^[A-Za-z0-9]')] #format for digits and letters but no symbols
		Quoting_Info = self.MOQ_Counting(Quoting_Info)

		#公斤裝公斤報價
		KG_Item = Quoting_Info["Item_Code"].astype(str).str.endswith("9")
		KG_Price = Quoting_Info["Price/M"]/Quoting_Info["Weight_Per_Pcs"]
		Roundup = KG_Price.apply(lambda x: np.ceil(x * 100) / 100 if pd.notna(x) else np.nan)
		Quoting_Info.loc[KG_Item, "Price/M"] = Roundup[KG_Item].map(lambda x: f"{x:.2f}/KG" if pd.notna(x) else "")

		#NQ的不要報
		status_col = Quoting_Info["Status"].fillna("").astype(str).str.strip().str.upper()
		Quoting_Info.loc[status_col.str.startswith("NQ"), "Price/M"] = None

		Quoting_Info.reset_index(drop=True, inplace=True)

		return Quoting_Info

	def Load_RFQ(self):
	
		for file in os.listdir(self.PATH):
			if file.startswith("600") and file.endswith(".xlsx"):
				RFQ_PATH = os.path.join(self.PATH, file)
		if 'RFQ_PATH' not in locals():
			raise FileNotFoundError("RFQ file starting with '600' not found.")

		DF_RFQ = pd.read_excel(RFQ_PATH, dtype=str, header=1)

		return DF_RFQ, RFQ_PATH
	
	def Fill_RFQ(self):

		Quoting_Info = self.Load_Cost_Info()
		DF_RFQ, RFQ_PATH = self.Load_RFQ()

		wb = load_workbook(RFQ_PATH) 
		ws = wb.active #activate C006 RFQ

		if "MOQ" not in DF_RFQ.columns:
			DF_RFQ.insert(11, "MOQ", None)
			ws["L2"] = "MOQ" #Add header

		for idx, row in DF_RFQ.iterrows():
			if row["Material"] in Quoting_Info["Item_Code"].values:
				found_item = Quoting_Info[Quoting_Info["Item_Code"] == row["Material"]].iloc[0]

				if not str(found_item["Status"]).startswith("NQ"):
					DF_RFQ.at[idx, "Price"] = found_item["Price/M"]
					DF_RFQ.at[idx, "MOQ"] = found_item["MOQ"]
				else:
					DF_RFQ.at[idx, "Price"] = ""  # Blank price for "NQ" status
			else:
				DF_RFQ.at[idx, "Price"] = ""  # Blank price for unmatched materials
		
		DF_RFQ["MOQ"] = DF_RFQ["MOQ"].replace(0, "")
		
		# Calculate column indices for writing to Excel
		price_col = DF_RFQ.columns.get_loc("Price") + 1
		moq_col = DF_RFQ.columns.get_loc("MOQ") + 1

		# Write updated DF_RFQ values back to Excel
		for idx, row in DF_RFQ.iterrows():
			ws.cell(row=idx+3, column=price_col).value = row["Price"]
			ws.cell(row=idx+3, column=moq_col).value = row["MOQ"]
		
		#修改部分標示為紅字
		red_font = Font(color="FF0000")

		for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
			row[price_col-1].font = red_font
			row[moq_col-1].font = red_font

		# Save the workbook (overwrite the original file)
		wb.save(RFQ_PATH)
		print("\nRFQ Filled, please check before quoting.")


class FILL_RFQ_C034(SHARED_FUNCTION):

	def __init__(self, path):
		super().__init__(path)
		self.PATH = path

	def Load_Cost_Info(self):

		Summarized_Sheet, DF_Cost_List = self.Concat_Cost_Sheet()

		Selected_Columns = [0, 1, 3, 4, 5, 6, 15, 51]
		Summarized_Sheet = Summarized_Sheet.iloc[:, Selected_Columns]
		Summarized_Sheet.iloc[:, 4:] = Summarized_Sheet.iloc[:, 4:].apply(pd.to_numeric, errors="coerce")

		Summarized_Sheet.rename(columns={
			Summarized_Sheet.columns[0]: "NO.",
			Summarized_Sheet.columns[1]: "Item_Code",
			Summarized_Sheet.columns[2]: "Size",
			Summarized_Sheet.columns[3]: "Status",
			Summarized_Sheet.columns[4]: "MOQ",
			Summarized_Sheet.columns[5]: "Quantity",
			Summarized_Sheet.columns[6]: "Weight_Per_Pcs",
			Summarized_Sheet.columns[7]: "Price/M",
		}, inplace = True)

		Quoting_Info = Summarized_Sheet[Summarized_Sheet["Item_Code"].astype(str).str.startswith("A")]

		Quoting_Info = self.MOQ_Counting(Quoting_Info)
		Quoting_Info.reset_index(drop=True, inplace=True)

		return Quoting_Info

	def Load_RFQ(self):

		RFQ_FILE = None
		header_row, start_row = None, None

		#判斷詢價單類別並讀取
		for file in os.listdir(self.PATH):
			if file.startswith("ABE") and file.endswith(".xlsx"):
				RFQ_FILE = os.path.join(self.PATH, file)
				header_row = 9
			elif file.startswith("TORNILLO") and file.endswith(".xlsx"):
				RFQ_FILE = os.path.join(self.PATH, file)
				header_row = 7

		if not RFQ_FILE:
			raise FileNotFoundError("No valid RFQ file found.")

		DF_RFQ = pd.read_excel(RFQ_FILE,  header=header_row, dtype=str)
		DF_RFQ = DF_RFQ.reset_index(drop=True)
		DF_RFQ["Price/M"] = None
		DF_RFQ["MOQ"] = None
		DF_RFQ = DF_RFQ[DF_RFQ["CODE"].str.startswith("A", na=False)]

		return DF_RFQ, RFQ_FILE, header_row

	def Fill_RFQ(self): 

		Quoting_Info = self.Load_Cost_Info()
		DF_RFQ, RFQ_FILE, header_row = self.Load_RFQ()

		wb = load_workbook(RFQ_FILE)
		ws = wb.active

		ws[f"H{header_row+1}"] = "Price/M"
		ws[f"I{header_row+1}"] = "MOQ"

		# Iterate through each row in DF_RFQ
		for idx, row in DF_RFQ.iterrows():
			if row["CODE"] in Quoting_Info["Item_Code"].values and row["QUANTITY PCS"] != 0:
				found_item = Quoting_Info[Quoting_Info["Item_Code"] == row["CODE"]].iloc[0]

				DF_RFQ.at[idx, "Price/M"] = found_item["Price/M"]
				DF_RFQ.at[idx, "MOQ"] = found_item["MOQ"]
			elif row["QUANTITY PCS"] != "0" and row["CODE"] not in Quoting_Info["Item_Code"].values:
				print(f"{row['CODE']} not found in 成本表")
			else:
				DF_RFQ.at[idx, "Price/M"] = "" 
				DF_RFQ.at[idx, "MOQ"] = "" 

		DF_RFQ["MOQ"] = DF_RFQ["MOQ"].replace(0, "")

		# Write updated DF_RFQ values back to Excel
		for idx, row in DF_RFQ.iterrows():
			ws.cell(row=header_row+2+idx, column=8).value = row["Price/M"] #ws 第一項為1非0 另除header再加一 共+2
			ws.cell(row=header_row+2+idx, column=9).value = row["MOQ"]
		   
		# Save the workbook (overwrite the original file)
		wb.save(RFQ_FILE)
		print("\nRFQ Filled, please check before quoting.")


class FILL_RFQ_D007(SHARED_FUNCTION):

	def __init__(self, path):
		super().__init__(path)
		self.PATH = path

	def Load_Cost_Info(self):

		Summarized_Sheet, DF_Cost_List = self.Concat_Cost_Sheet()

		Selected_Columns = [0, 1, 3, 4, 5, 6, 15, 52]
		Summarized_Sheet = Summarized_Sheet.iloc[:, Selected_Columns]

		Summarized_Sheet.iloc[:, 4:] = Summarized_Sheet.iloc[:, 4:].apply(pd.to_numeric, errors="coerce")

		Summarized_Sheet.rename(columns={
			Summarized_Sheet.columns[0]: "NO.",
			Summarized_Sheet.columns[1]: "Item_Code",
			Summarized_Sheet.columns[2]: "Size",
			Summarized_Sheet.columns[3]: "Status",
			Summarized_Sheet.columns[4]: "MOQ",
			Summarized_Sheet.columns[5]: "Quantity",
			Summarized_Sheet.columns[6]: "Weight_Per_Pcs",
			Summarized_Sheet.columns[7]: "Price/CTN",
		}, inplace = True)

		# Create a clean data named Quoting_Info, arrange according to RFQ item no.
		Quoting_Info = Summarized_Sheet[Summarized_Sheet["Item_Code"].astype(str).str.startswith("H")] 

		status_col = Quoting_Info["Status"].fillna("").astype(str).str.strip().str.upper()
		Quoting_Info.loc[status_col.str.startswith("NQ"), "Price/CTN"] = None

		Quoting_Info.reset_index(drop=True, inplace=True)

		return Quoting_Info
	
	def Load_RFQ(self):

		for file in os.listdir(self.PATH):
			if "RFQ" in file and file.endswith(".xlsx"):
				RFQ_FILE = os.path.join(self.PATH, file)

		DF_RFQ = pd.read_excel(RFQ_FILE, skiprows=2, dtype=str)

		return DF_RFQ, RFQ_FILE

	def Fill_RFQ(self):

		Quoting_Info = self.Load_Cost_Info()
		DF_RFQ, RFQ_FILE = self.Load_RFQ()

		# Iterate through each row in DF_RFQ
		for idx, row in DF_RFQ.iterrows():
			# Check if the "Material" in DF_RFQ exists in Quoting_Info["Item_Code"]
			if row["Huttig SKU"] in Quoting_Info["Item_Code"].values:
				# Find the corresponding index in Quoting_Info
				find_item = Quoting_Info[Quoting_Info["Item_Code"] == row["Huttig SKU"]].index
		 
				# Use iloc with integer column indices
				DF_RFQ.iloc[idx, DF_RFQ.columns.get_loc("FOB Origin")] = Quoting_Info.iloc[find_item[0]]["Price/CTN"]
			else:
				DF_RFQ.iloc[idx, DF_RFQ.columns.get_loc("FOB Origin")] = None

		DF_RFQ.to_excel(os.path.join(self.PATH, "Quoting.xlsx"), index=False)
		print("\nFrom Quoting.xlsx Copy and Paste 'FOB Origin' to RFQ.")
		return
		
		#覆寫D007文件後無法修改，且不確定能不能上傳系統系統故在此步驟前return			 
		wb = load_workbook(RFQ_FILE)
		ws = wb.active 

		for idx, row in DF_RFQ.iterrows():
			price_col = DF_RFQ.columns.get_loc("FOB Origin") + 1

			# Define a function to ensure the value is scalar
			def get_scalar_value(value):
				if isinstance(value, (pd.Series, np.ndarray)):
					return value.item()  # Return the scalar value if it's a Series or ndarray
				return value  # Return the value directly if it's already a scalar

			# Write the values to Excel using the get_scalar_value function 
			ws.cell(row=idx + 4, column=price_col).value = get_scalar_value(row["FOB Origin"])
		   
		wb.save(RFQ_FILE)
		print("\nRFQ Filled, please check before quoting.")


class FILL_RFQ_C019(SHARED_FUNCTION):
	
	def __init__(self, path):
		super().__init__(path)
		self.PATH = path

	def Load_Cost_Info(self):

		Summarized_Sheet, DF_Cost_List = self.Concat_Cost_Sheet()

		selected_columns = [
		"Unnamed: 0",     # NO.
		"客戶代號",        # Item_Code
		"孔穴牙長",        # Status
		"尺寸",           
		"MOQ",           # MOQ
		"數量(M)",        # Quantity
		"加工價(單重/M)",  # Weight_Per_Pcs
		"總計/M",         # Price
		"相關費用"         # Tooling Fee
		]

		# Step 2: Rename them
		rename_mapping = {
			"Unnamed: 0": "NO.",
			"客戶代號": "Item_Code",
			"孔穴牙長": "Status",
			"尺寸": "Size",
			"MOQ": "MOQ",
			"數量(M)": "Quantity",
			"加工價(單重/M)": "Weight_Per_Pcs",
			"總計/M": "Price",
			"相關費用": "Tooling Fee"
		}
		
		# Step 3: Apply selection and renaming
		Quoting_Info = Summarized_Sheet[selected_columns].rename(columns=rename_mapping)
		Quoting_Info.iloc[:, 4:] = Quoting_Info.iloc[:, 4:].apply(pd.to_numeric, errors="coerce")

		# Create a clean data named Quoting_Info, arrange according to RFQ item no.
		Quoting_Info = Quoting_Info[Quoting_Info["Item_Code"].astype(str).str.len() == 21]
		# Quoting_Info = Quoting_Info.sort_values(by="NO.", ascending=True)

		#Price Counting M to H
		Quoting_Info["Price/H"] = Quoting_Info["Price"].apply(lambda x: np.ceil(x/10*100)/100 if pd.notna(x) else np.nan)

		Quoting_Info["MOQ"] = 300/Quoting_Info["Weight_Per_Pcs"]
		Quoting_Info["Length"] = Quoting_Info["Item_Code"].apply(lambda x: ".".join(x.split(".")[3:]))
		Quoting_Info.loc[(Quoting_Info["Length"].astype(float) <= 100) & (Quoting_Info["MOQ"] < 100), "MOQ"] = 100 
		Quoting_Info.loc[(Quoting_Info["Length"].astype(float) > 100) & (Quoting_Info["MOQ"] < 50), "MOQ"] = 50
		# Roundup MOQ
		Quoting_Info["MOQ"] = Quoting_Info["MOQ"].apply(lambda x: int(np.ceil(x / 10) * 10) if pd.notna(x) else np.nan)
		# Add new column name for MOQ Comment
		Quoting_Info["MOQ_Comment"] = " "
		Quoting_Info.loc[Quoting_Info["MOQ"] > Quoting_Info["Quantity"], "MOQ_Comment"] = \
		   Quoting_Info.loc[Quoting_Info["MOQ"] > Quoting_Info["Quantity"], "MOQ"].apply(
				lambda x: f"MOQ: {int(x)}M" if pd.notna(x) else " ")
		
		# Tooling Fee Counting
		Quoting_Info.loc[Quoting_Info["Tooling Fee"] == 0, "Tooling Fee"] = None
			
		Quoting_Info["Tooling Fee"] = Quoting_Info["Tooling Fee"].apply(
			lambda x: int(np.ceil(x / 30 /10) * 10) if pd.notna(x) else np.nan)
		
		# Define the function to check if Tooling Fee is not NaN and not an empty string
		def check_tooling_fee(value):
			return pd.notna(value) and value != " "   
		# Add new column name for Tooling Fee Comment
		Quoting_Info["Tooling_Fee_Comment"] = " "
		tooling_fee_condition = Quoting_Info["Tooling Fee"].apply(check_tooling_fee)
		Quoting_Info.loc[tooling_fee_condition, "Tooling_Fee_Comment"]= \
			Quoting_Info.loc[tooling_fee_condition, "Tooling Fee"].apply(
				lambda x: f"Tooling Fee: {int(x)} EURO" if pd.notna(x) else "")

		# Comment that combines MOQ and Tooling Fee info for filling up RFQ
		Quoting_Info["Comment"] = Quoting_Info["MOQ_Comment"] + "\n" + Quoting_Info["Tooling_Fee_Comment"]

		# If MOQ > Quantity, revise quantity with its MOQ
		Quoting_Info.loc[Quoting_Info["MOQ"] > Quoting_Info["Quantity"], "Quantity"] = \
			Quoting_Info.loc[Quoting_Info["MOQ"] > Quoting_Info["Quantity"], "MOQ"]

		# switch quantity from M to pcs for later filling up RFQ
		Quoting_Info["Quantity"] = (Quoting_Info["Quantity"] * 1000).astype(int)

		#If Not Quoting, remain the price as 0.00 and reject
		status_str = Quoting_Info["Status"].fillna("").astype(str)
		Quoting_Info.loc[status_str.str.startswith("NQ"), "Price"] = "0.00"
		Quoting_Info.loc[status_str.str.startswith("NQ"), "Price/H"] = "0.00"
		Quoting_Info.loc[status_str.str.startswith("NQ"), "Quot. Comment"] = "AB2-Rejection for quality reasons"
		Quoting_Info.loc[status_str.str.startswith("NQ"), "Comment"] = " "

		return Quoting_Info

	def Load_RFQ(self):
		Quoting_Info = self.Load_Cost_Info()
		
		for file in os.listdir(self.PATH):
			if "_RFQ_" in file and file.endswith(".xlsx"):
				DF_RFQ = os.path.join(self.PATH, file)
				DF_RFQ = pd.read_excel(DF_RFQ, header=9, sheet_name="RFQ positions", dtype=str)

		DF_RFQ["Price/H"] = None

		# Iterate through each row in DF_RFQ
		for idx, row in DF_RFQ.iterrows():
			# Check if the "Material" in DF_RFQ exists in Quoting_Info["Item_Code"]
			if row["Material"] in Quoting_Info["Item_Code"].values:
				# Find the corresponding index in Quoting_Info
				find_item = Quoting_Info[Quoting_Info["Item_Code"] == row["Material"]]
				
				# Use .loc with column names for assignment
				DF_RFQ.loc[idx, "RFQ Quantity"] = find_item["Quantity"].values[0]
				DF_RFQ.loc[idx, "Price"] = find_item["Price"].values[0]
				DF_RFQ.loc[idx, "Price/H"] = find_item["Price/H"].values[0]
				DF_RFQ.loc[idx, "Quot. Comment"] = find_item["Quot. Comment"].values[0]
				DF_RFQ.loc[idx, "Comment"] = find_item["Comment"].values[0]
			else:
				DF_RFQ.loc[idx, "Price"] = "0.00"
				DF_RFQ.loc[idx, "Price/H"] = "0.00"
				DF_RFQ.loc[idx, "Quot. Comment"] = "AB2-Rejection for quality reasons"
				DF_RFQ.loc[idx, "Comment"] = ""
				print(f"Item {row['Item']}   {row['Material']} not found in 成本表.")

		filled_columns = ["Item", "Material", "RFQ Quantity", "Price", "Price/H", "Quot. Comment","Comment"]
		output_path = os.path.join(self.PATH, "Quoting.xlsx")
		DF_RFQ.to_excel(output_path, columns=filled_columns, index=False)

		return DF_RFQ, output_path

	#C019詢價單直接填入後會被系統拒絕，故以輸出檔案整欄貼上。
	def Fill_RFQ(self):
		
		Quoting_Info = self.Load_Cost_Info()
		DF_RFQ, output_path = self.Load_RFQ()

		wb = load_workbook(output_path)
		ws = wb.active

		red_font = Font(color="FF0000")
		ws["I2"] = "Copy and Paste Column: RFQ Quantity, Price, Quot. Comment, Comment"
		ws["I3"] = "!Remember to fill price unit and del. time!"
		ws["I2"].font = red_font
		ws["I3"].font = red_font

		wb.save(output_path)

		print("\nQuoting information available for pasting.")
		print("Check Before Quoting!")


if __name__ == "__main__":

	path_input = input(r"Enter File Path: ") 
	customer = input("Enter Customer Code: ").upper() 
	
	expected_class = f"FILL_RFQ_{customer}"
	class_name = [name for name in globals() if name.upper() == expected_class]

	if class_name:
		class_name = class_name[0]
		class_ref = globals()[class_name] # Get the class object
		instance = class_ref(path_input)  # Instantiate the class
		instance.Fill_RFQ()               # Call the method                          
		instance.Count_Quote_Weight()     # Call the method
		
	else:
		print(f"Error: No class found for customer code {customer}. Please check your input.")
