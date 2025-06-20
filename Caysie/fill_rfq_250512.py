import sys
import pandas as pd
import os
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font

sys.path.append(r"C:\Users\sales6\Desktop\Work")
from shared_function import Concat_Cost_Sheet
from shared_function import MOQ_Counting

def Fill_RFQ_C006(FILE_PATH):

	for root, dirs, files in os.walk(FILE_PATH):
		for file in files:
			if file.startswith("600") and file.endswith(".xlsx"):
				RFQ_PATH = os.path.join(root, file)

	Summarized_Sheet = Concat_Cost_Sheet(FILE_PATH)

	Selected_Columns = [
		Summarized_Sheet.columns[0], 
		"客戶代號",
		"尺寸",
		"孔穴牙長",
		"MOQ",
		"數量(M)",
		"加工價(單重/M)",
		"總計/M"]
	
	# Filter the Summarized_Sheet to keep only the selected columns
	Summarized_Sheet = Summarized_Sheet[Selected_Columns]
	
	Summarized_Sheet.iloc[:, 4:] = Summarized_Sheet.iloc[:, 4:].apply(pd.to_numeric, errors="coerce")

	Summarized_Sheet.rename(columns={
		Summarized_Sheet.columns[0]: "NO.",
		Summarized_Sheet.columns[1]: "Item_Code",
		Summarized_Sheet.columns[2]: "Size",
		Summarized_Sheet.columns[3]: "Status",
		Summarized_Sheet.columns[4]: "MOQ",
		Summarized_Sheet.columns[5]: "Quantity",
		Summarized_Sheet.columns[6]: "Weight_Per_Pcs",
		Summarized_Sheet.columns[7]: "Price",
	}, inplace = True)

	# Create a clean data named Quoting_Info, arrange according to RFQ item no.
	Quoting_Info = Summarized_Sheet[Summarized_Sheet["Item_Code"].astype(str).str.match(r'^[A-Za-z0-9]')]
	Quoting_Info = MOQ_Counting(Quoting_Info)
	Quoting_Info.reset_index(drop=True, inplace=True)

	# Find C019 RFQ File
	DF_RFQ = pd.read_excel(RFQ_PATH, dtype=str)
	DF_RFQ.insert(11, "MOQ", None) #Insert column "MOQ" at column 11 set as none

	# Iterate through each row in DF_RFQ
	for idx, row in DF_RFQ.iterrows():
		# Check if the "Material" in DF_RFQ exists in Quoting_Info["Item_Code"]
		if row["Material"] in Quoting_Info["Item_Code"].values:
			# Find the corresponding index in Quoting_Info
			find_item = Quoting_Info[Quoting_Info["Item_Code"] == row["Material"]].index
	 
			# Use iloc with integer column indices
			DF_RFQ.iloc[idx, DF_RFQ.columns.get_loc("Price")] = Quoting_Info.iloc[find_item[0]]["Price"]
			DF_RFQ.iloc[idx, DF_RFQ.columns.get_loc("MOQ")] = int(Quoting_Info.iloc[find_item[0]]["MOQ"])
		else:
			pass
			
	DF_RFQ["MOQ"] = DF_RFQ["MOQ"].replace(0, "")

	# Load the Excel file using openpyxl to preserve formatting
	wb = load_workbook(RFQ_PATH)
	ws = wb.active 

	ws["L1"] = "MOQ"

	# After updating DF_RFQ, write it back into the Excel file while preserving the formatting
	for idx, row in DF_RFQ.iterrows():
		# Assuming that these columns are in specific columns in the sheet
		# Find the columns that might require altering and get values
		price_col = DF_RFQ.columns.get_loc("Price") + 1
		moq_col = DF_RFQ.columns.get_loc("MOQ") + 1
	  
		# Define a function to ensure the value is scalar
		def get_scalar_value(value):
			if isinstance(value, (pd.Series, np.ndarray)):
				return value.item()  # Return the scalar value if it's a Series or ndarray
			return value  # Return the value directly if it's already a scalar

		# Write the values to Excel using the helper function
		# Filling the RFQ columns with  
		ws.cell(row=idx+2, column=price_col).value = get_scalar_value(row["Price"])
		ws.cell(row=idx+2, column=moq_col).value = get_scalar_value(row["MOQ"])
	
	red_font = Font(color="FF0000")

	for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=price_col, max_col=price_col):
		for cell in row:
			cell.font = red_font

	for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=moq_col, max_col=moq_col):
		for cell in row:
			cell.font = red_font

	# Save the workbook (overwrite the original file)
	wb.save(RFQ_PATH)




def Fill_RFQ_C034(FILE_PATH):

	RFQ_FILE = None
	header_row, start_row = None, None  # Default values

	for root, dirs, files in os.walk(FILE_PATH):
		for file in files:
			if file.startswith("ABE") and file.endswith(".xlsx"):
				RFQ_FILE = os.path.join(root, file)
				header_row, start_row = 9, 11
			elif file.startswith("TORNILLO") and file.endswith(".xlsx"):
				RFQ_FILE = os.path.join(root, file)
				header_row, start_row = 7, 9

	if not RFQ_FILE:
		raise FileNotFoundError("No valid RFQ file found.")

	Summarized_Sheet = Concat_Cost_Sheet(FILE_PATH)

	Selected_Columns = [0, 1, 3, 4, 5, 6, 15, 51]
	Summarized_Sheet = Summarized_Sheet.iloc[:, Selected_Columns]
	Summarized_Sheet.iloc[:, 4:] = Summarized_Sheet.iloc[:, 4:].apply(pd.to_numeric, errors="coerce")

	Summarized_Sheet.rename(columns={
		Summarized_Sheet.columns[0]: "NO.",
		Summarized_Sheet.columns[1]: "Item_Code",
		Summarized_Sheet.columns[2]: "Size",
		Summarized_Sheet.columns[3]: "Status",
		Summarized_Sheet.columns[4]: "MOQ",
		Summarized_Sheet.columns[5]: "Quantity",
		Summarized_Sheet.columns[6]: "Weight_Per_Pcs",
		Summarized_Sheet.columns[7]: "Price",
	}, inplace = True)

	Quoting_Info = Summarized_Sheet[Summarized_Sheet["Item_Code"].astype(str).str.startswith("A")]

	Quoting_Info = MOQ_Counting(Quoting_Info)
	Quoting_Info.reset_index(drop=True, inplace=True)

#注意C034的skiprow or header 行數不同

	DF_RFQ = pd.read_excel(RFQ_FILE,  header=header_row, dtype=str)

	DF_RFQ["Price/M"] = None
	DF_RFQ["MOQ"] = None

	# Iterate through each row in DF_RFQ
	for idx, row in DF_RFQ.iterrows():
		# Check if the "Material" in DF_RFQ exists in Quoting_Info["Item_Code"]
		if row["CODE"] in Quoting_Info["Item_Code"].values and row["QUANTITY PCS"] != "-":
			# Find the corresponding index in Quoting_Info
			find_item = Quoting_Info[Quoting_Info["Item_Code"] == row["CODE"]].index

			price = Quoting_Info.loc[find_item[0], "Price"]
			moq_val = int(Quoting_Info.loc[find_item[0], "MOQ"])

			# Assign values to DF_RFQ, converting MOQ 0 to None
			DF_RFQ.iloc[idx, DF_RFQ.columns.get_loc("Price/M")] = price
			DF_RFQ.iloc[idx, DF_RFQ.columns.get_loc("MOQ")] = None if moq_val == 0 else moq_val
	 
		else:
			pass
			
	# Load the Excel file using openpyxl to preserve formatting
	wb = load_workbook(RFQ_FILE)
	ws = wb.active 

	ws.cell(row=header_row+1, column=8, value="Price/M")
	ws.cell(row=header_row+1, column=9, value="MOQ")

	# After updating DF_RFQ, write it back into the Excel file while preserving the formatting
	for idx, row in DF_RFQ.iterrows():
		# Assuming that these columns are in specific columns in the sheet
		# Find the columns that might require altering and get values
		price_col = DF_RFQ.columns.get_loc("Price/M") + 1
		moq_col = DF_RFQ.columns.get_loc("MOQ") + 1
	  
		# Define a function to ensure the value is scalar
		def get_scalar_value(value):
			if isinstance(value, (pd.Series, np.ndarray)):
				return value.item()  # Return the scalar value if it's a Series or ndarray
			return value  # Return the value directly if it's already a scalar

		# Write the values to Excel using the helper function
		# Filling the RFQ columns with  
		ws.cell(row=idx + start_row, column=price_col).value = get_scalar_value(row["Price/M"])
		ws.cell(row=idx + start_row, column=moq_col).value = get_scalar_value(row["MOQ"])
	   
	# Save the workbook (overwrite the original file)
	wb.save(RFQ_FILE)



def Fill_RFQ_D007(FILE_PATH):

	print("此功能尚未修復，請勿使用。")

	Summarized_Sheet = Concat_Cost_Sheet(FILE_PATH)

	Selected_Columns = [0, 1, 3, 4, 5, 6, 15, 52]
	Summarized_Sheet = Summarized_Sheet.iloc[:, Selected_Columns]

	Summarized_Sheet.iloc[:, 3:] = Summarized_Sheet.iloc[:, 3:].apply(pd.to_numeric, errors="coerce")

	Summarized_Sheet.rename(columns={
		Summarized_Sheet.columns[0]: "NO.",
		Summarized_Sheet.columns[1]: "Item_Code",
		Summarized_Sheet.columns[2]: "Size",
		Summarized_Sheet.columns[3]: "Status",
		Summarized_Sheet.columns[4]: "MOQ",
		Summarized_Sheet.columns[5]: "Quantity",
		Summarized_Sheet.columns[6]: "Weight_Per_Pcs",
		Summarized_Sheet.columns[7]: "Price/CTN",
	}, inplace = True)

	# Create a clean data named Quoting_Info, arrange according to RFQ item no.
	Quoting_Info = Summarized_Sheet[Summarized_Sheet["Item_Code"].astype(str).str.startswith("H")]

	Quoting_Info.reset_index(drop=True, inplace=True)

	for root, dirs, files in os.walk(FILE_PATH):
		for file in files:
			if "RFQ" in file and file.endswith(".xlsx"):
				RFQ_FILE = os.path.join(root, file)

	DF_RFQ = pd.read_excel(RFQ_FILE, skiprows=2, dtype=str)

	# Iterate through each row in DF_RFQ
	for idx, row in DF_RFQ.iterrows():
		# Check if the "Material" in DF_RFQ exists in Quoting_Info["Item_Code"]
		if row["Huttig SKU"] in Quoting_Info["Item_Code"].values:
			# Find the corresponding index in Quoting_Info
			find_item = Quoting_Info[Quoting_Info["Item_Code"] == row["Huttig SKU"]].index
	 
			# Use iloc with integer column indices
			DF_RFQ.iloc[idx, DF_RFQ.columns.get_loc("FOB Origin")] = Quoting_Info.iloc[find_item[0]]["Price/CTN"]
		else:
			DF_RFQ.iloc[idx, DF_RFQ.columns.get_loc("FOB Origin")] = None
				 
	# Load the Excel file using openpyxl to preserve formatting
	wb = load_workbook(RFQ_FILE)
	ws = wb.active 

	# After updating DF_RFQ, write it back into the Excel file while preserving the formatting
	for idx, row in DF_RFQ.iterrows():
		# Assuming that these columns are in specific columns in the sheet
		# Find the columns that might require altering and get values
		price_col = DF_RFQ.columns.get_loc("FOB Origin") + 1


		# Define a function to ensure the value is scalar
		def get_scalar_value(value):
			if isinstance(value, (pd.Series, np.ndarray)):
				return value.item()  # Return the scalar value if it's a Series or ndarray
			return value  # Return the value directly if it's already a scalar

		# Write the values to Excel using the helper function
		# Filling the RFQ columns with  
		ws.cell(row=idx + 4, column=price_col).value = get_scalar_value(row["FOB Origin"])
	   
	# Save the workbook (overwrite the original file)
	wb.save(RFQ_FILE)


def Fill_RFQ_C019(FILE_PATH):

	Summarized_Sheet = Concat_Cost_Sheet(FILE_PATH)

	selected_columns = [
	"Unnamed: 0",     # NO.
	"客戶代號",        # Item_Code
	"孔穴牙長",        # Status
	"尺寸",           
	"MOQ",           # MOQ
	"數量(M)",        # Quantity
	"加工價(單重/M)",  # Weight_Per_Pcs
	"總計/M",         # Price
	"相關費用"         # Tooling Fee
	]

	# Step 2: Rename them
	rename_mapping = {
		"Unnamed: 0": "NO.",
		"客戶代號": "Item_Code",
		"孔穴牙長": "Status",
		"尺寸": "Size",
		"MOQ": "MOQ",
		"數量(M)": "Quantity",
		"加工價(單重/M)": "Weight_Per_Pcs",
		"總計/M": "Price",
		"相關費用": "Tooling Fee"
	}

	
	# Step 3: Apply selection and renaming
	Quoting_Info = Summarized_Sheet[selected_columns].rename(columns=rename_mapping)
	Quoting_Info.iloc[:, 4:] = Quoting_Info.iloc[:, 4:].apply(pd.to_numeric, errors="coerce")

	# Create a clean data named Quoting_Info, arrange according to RFQ item no.
	Quoting_Info = Quoting_Info[Quoting_Info["Item_Code"].astype(str).str.len() == 21]
	# Quoting_Info = Quoting_Info.sort_values(by="NO.", ascending=True)

	#Price Counting M to H
	Quoting_Info["Price/H"] = Quoting_Info["Price"].apply(lambda x: np.ceil(x/10*100)/100 if pd.notna(x) else np.nan)

	Quoting_Info["MOQ"] = 300/Quoting_Info["Weight_Per_Pcs"]
	Quoting_Info["Length"] = Quoting_Info["Item_Code"].apply(lambda x: ".".join(x.split(".")[3:]))
	Quoting_Info.loc[(Quoting_Info["Length"].astype(float) <= 100) & (Quoting_Info["MOQ"] < 100), "MOQ"] = 100 
	Quoting_Info.loc[(Quoting_Info["Length"].astype(float) > 100) & (Quoting_Info["MOQ"] < 50), "MOQ"] = 50
	# Roundup MOQ
	Quoting_Info["MOQ"] = Quoting_Info["MOQ"].apply(lambda x: int(np.ceil(x / 10) * 10) if pd.notna(x) else np.nan)
	# Add new column name for MOQ Comment
	Quoting_Info["MOQ_Comment"] = " "
	Quoting_Info.loc[Quoting_Info["MOQ"] > Quoting_Info["Quantity"], "MOQ_Comment"] = \
	   Quoting_Info.loc[Quoting_Info["MOQ"] > Quoting_Info["Quantity"], "MOQ"].apply(
	        lambda x: f"MOQ: {int(x)}M" if pd.notna(x) else " ")
	
	# Tooling Fee Counting
	Quoting_Info.loc[Quoting_Info["Tooling Fee"] == 0, "Tooling Fee"] = None
		
	Quoting_Info["Tooling Fee"] = Quoting_Info["Tooling Fee"].apply(
		lambda x: int(np.ceil(x / 30 /10) * 10) if pd.notna(x) else np.nan)
	# Define the function to check if Tooling Fee is not NaN and not an empty string
	
	def check_tooling_fee(value):
		return pd.notna(value) and value != " "   
	# Add new column name for Tooling Fee Comment
	Quoting_Info["Tooling_Fee_Comment"] = " "
	tooling_fee_condition = Quoting_Info["Tooling Fee"].apply(check_tooling_fee)
	Quoting_Info.loc[tooling_fee_condition, "Tooling_Fee_Comment"]= \
		Quoting_Info.loc[tooling_fee_condition, "Tooling Fee"].apply(
			lambda x: f"Tooling Fee: {int(x)} EURO" if pd.notna(x) else "")

	# Comment that combines MOQ and Tooling Fee info for filling up RFQ
	Quoting_Info["Comment"] = Quoting_Info["MOQ_Comment"] + "\n" + Quoting_Info["Tooling_Fee_Comment"]


	# switch quantity from M to pcs for later filling up RFQ
	

	#If Not Quoting, remain the price as 0.00 and reject
	# AB4 used to be: not in production range, awaiting for C019 Portal fixing.
	status_str = Quoting_Info["Status"].fillna("").astype(str)
	Quoting_Info.loc[status_str.str.startswith("NQ"), "Price"] = "0.00"
	Quoting_Info.loc[status_str.str.startswith("NQ"), "Quot. Comment"] = "AB2-Rejection for quality reasons"
	Quoting_Info.loc[status_str.str.startswith("NQ"), "Comment"] = " "


	#get total weight for filling 詢價統計表
	Total_Weight= (Quoting_Info["Weight_Per_Pcs"]*Quoting_Info["Quantity"]/1000).sum()
	print(f"\n詢價總重 {Total_Weight} kg\n")

	for root, dirs, files in os.walk(FILE_PATH):
		for file in files:
			if "_RFQ_" in file and file.endswith(".xlsx"):
				DF_RFQ = os.path.join(root, file)
				DF_RFQ = pd.read_excel(DF_RFQ, header=9, sheet_name="RFQ positions", dtype=str)

	DF_RFQ["Price/H"] = None

	# Iterate through each row in DF_RFQ
	for idx, row in DF_RFQ.iterrows():
		# Check if the "Material" in DF_RFQ exists in Quoting_Info["Item_Code"]
		if row["Material"] in Quoting_Info["Item_Code"].values:
			# Find the corresponding index in Quoting_Info
			find_item = Quoting_Info[Quoting_Info["Item_Code"] == row["Material"]]
			
			# Use .loc with column names for assignment
			DF_RFQ.loc[idx, "RFQ Quantity"] = find_item["Quantity"].values[0]
			DF_RFQ.loc[idx, "Price"] = find_item["Price"].values[0]
			DF_RFQ.loc[idx, "Price/H"] = find_item["Price/H"].values[0]
			DF_RFQ.loc[idx, "Quot. Comment"] = find_item["Quot. Comment"].values[0]
			DF_RFQ.loc[idx, "Comment"] = find_item["Comment"].values[0]
		else:
			DF_RFQ.loc[idx, "Price"] = "0.00"
			DF_RFQ.loc[idx, "Price/H"] = "0.00"
			DF_RFQ.loc[idx, "Quot. Comment"] = "AB2-Rejection for quality reasons"
			DF_RFQ.loc[idx, "Comment"] = ""
			print(f"Item {row['Item']}   {row['Material']} not found in 成本表.")

	filled_columns = ["Item", "Material", "RFQ Quantity", "Price", "Price/H", "Quot. Comment","Comment"]
	output_path = os.path.join(FILE_PATH, "Quoting.xlsx")
	DF_RFQ.to_excel(output_path, columns=filled_columns, index=False)
	print("\nQuoting information available for pasting.")
	print("!Please check price unit and del. time!")




if __name__ == "__main__":

	path_input = input(r"Enter File Path: ") 
	customer = input("Enter Customer Code: ")  # Fixed variable name
	
	function_name = f"Fill_RFQ_{customer}"

	if function_name in globals():
		fill_function = globals()[function_name]  # Corrected indentation
		fill = fill_function(path_input)
		print("RFQ Filled, please check before quoting.")
	else:
		print(f"Error: No function found for customer code {customer}. Please check your input.")
