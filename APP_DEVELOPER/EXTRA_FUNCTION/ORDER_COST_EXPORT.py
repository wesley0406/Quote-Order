import pandas as pd
import os
from datetime import datetime
import cx_Oracle
import re
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

class ORDER_COST_EXPORT_CLS() : 
	def __init__(self, path, EX_RATE) : 
		self.ALL = None
		self.Item_List = None
		self.SC_Number = None
		self.Order_Number = None
		self.Customer_Code = None
		self.PATH = path
		self.EXCHANGE = float(EX_RATE)
		self.ORDER_INFO()
		self.FETCH_DATA()
		self.CALCULATION()
		self.FILL_EXCEL_EXPORT()

	def ORDER_INFO(self):
		SSL_excel = None
		for file in os.listdir(self.PATH):
			print(file)
			if "0210M" in file:
				SSL_excel = os.path.join(self.PATH, file)
		if not SSL_excel: 
			raise ValueError("Please download the SSL0210 excel file")
			exit()
		else :
			self.ALL = pd.read_excel(SSL_excel, dtype={'客戶產品代號(P/N)': str})

	def FETCH_DATA(self):
		First_Item = self.ALL["客戶產品代號(P/N)"].iloc[0]
		# Define connection details
		dsn = cx_Oracle.makedsn(
			host = "192.168.1.242",      # Replace with your host IP or domain
			port = 1526,                # Replace with your port
			service_name = "sperpdb"  # Replace with your service name
		)

		# Establish the connection
		connection = cx_Oracle.connect(
			user = "spselect",         # Replace with your username
			password = "select",     # Replace with your password
			dsn = dsn
		)
		query1 = f"SELECT SC_NO,CST_REFE_NO,ORD_CST_NO FROM V_SCH0200Q_ORD WHERE CST_PART_NO = '{First_Item}' ORDER BY SC_NO DESC"
		df1 = pd.read_sql_query(query1, connection)
		self.SC_Number = df1["SC_NO"].iloc[0]
		self.Order_Number = df1["CST_REFE_NO"].iloc[0]
		self.Customer_Code = df1["ORD_CST_NO"].iloc[0]
		
		query2 = f"SELECT * FROM ssl_cst_orde_d WHERE SC_NO = '{self.SC_Number}' "
		self.Item_List = pd.read_sql_query(query2, connection)

	def CALCULATION(self):
		self.Item_List["線徑"] = round(self.ALL["線徑"],2)
		self.Item_List["規格"] = self.ALL["規格"]
		self.Item_List["訂單總箱數"] = self.Item_List["KEGS"].sum()
		self.Item_List["訂單總KGS"] = self.Item_List["ORDER_WEIG"].sum()
		self.Item_List["訂單總M數"] = self.Item_List["ORDER_QTY"].sum()
		self.Item_List["訂單總金額"] = self.Item_List["ORDER_AMT"].sum()
		self.Item_List["訂單總成本(NTD)"] = self.Item_List["COST_AMT"].sum()
		self.Item_List["匯率"] = float(self.EXCHANGE)
		self.Item_List["訂單總毛利(原幣)"] = self.Item_List["訂單總金額"]-(self.Item_List["訂單總成本(NTD)"]/self.Item_List["匯率"])
		self.Item_List["訂單利潤率"] = (self.Item_List["訂單總金額"]/(self.Item_List["訂單總成本(NTD)"]/self.Item_List["匯率"])-1)*100
		self.Item_List["利潤率"] = (self.Item_List["ORDER_AMT"]*self.Item_List["匯率"]/self.Item_List["COST_AMT"]-1)*100
		
	def FILL_EXCEL_EXPORT(self):
		wb = Workbook()
		ws = wb.active

		ws["A1"] = "訂單成本表"
		ws["A2"] = "客戶代號"
		ws["B2"] = self.Customer_Code
		ws["A3"] = "SC"
		ws["B3"] = self.SC_Number
		ws["A4"] = "客戶訂單號碼"
		ws["B4"] = self.Order_Number


		headers = [
			"項次", "客戶產品代號(P/N)", "線徑", "棧板箱數", "計量單位", "箱數", "棧板數",
			"訂單項次KGS", "訂單項次M數", "訂單單價", "訂單金額(原幣)", "成本單價(NTD/M)",
			"成本金額(NTD)", "生管交期", "訂單總箱數", "訂單總KGS", "訂單總M數", "訂單總金額",
			"規格", "報價千支重", "訂單總成本(NTD)", "訂單總毛利(原幣)", "訂單利潤率", "利潤率",
			"線材單價(NTD/KG)", "匯率"
		]

		for col, header in enumerate(headers, start=1):  # start=1 for column A
			ws.cell(row=5, column=col, value=header)

		header_row = 6  # Start writing data from row 8 (just below the headers)


		for i, row in self.Item_List.iterrows():
			ws[f"A{header_row + i}"] = row["KIND_NO"]
			ws[f"B{header_row + i}"] = row["CST_PART_NO"]
			ws[f"C{header_row + i}"] = row["線徑"]
			ws[f"D{header_row + i}"] = row["CTN_PLT"]
			ws[f"E{header_row + i}"] = row["UNIT_NAME"]
			ws[f"F{header_row + i}"] = row["KEGS"]
			ws[f"G{header_row + i}"] = row["PLT_QTY"]
			ws[f"H{header_row + i}"] = row["ORDER_WEIG"]
			ws[f"I{header_row + i}"] = row["ORDER_QTY"]
			ws[f"J{header_row + i}"] = row["PRICE"]
			ws[f"K{header_row + i}"] = row["ORDER_AMT"]
			ws[f"L{header_row + i}"] = row["COST_PRICE"]
			ws[f"M{header_row + i}"] = row["COST_AMT"]
			ws[f"N{header_row + i}"] = row["VEN_DLV_DATE"].strftime("%Y/%m/%d")
			ws[f"O{header_row + i}"] = row["訂單總箱數"]
			ws[f"P{header_row + i}"] = row["訂單總KGS"]
			ws[f"Q{header_row + i}"] = row["訂單總M數"]
			ws[f"R{header_row + i}"] = row["訂單總金額"]
			ws[f"S{header_row + i}"] = row["規格"]
			ws[f"T{header_row + i}"] = row["PDC_1000_WT"]
			ws[f"U{header_row + i}"] = row["訂單總成本(NTD)"]
			ws[f"V{header_row + i}"] = row["訂單總毛利(原幣)"]
			ws[f"W{header_row + i}"] = row["訂單利潤率"]
			ws[f"X{header_row + i}"] = row["利潤率"]
			ws[f"Y{header_row + i}"] = row["DRW_PRICE"]
			ws[f"Z{header_row + i}"] = row["匯率"]

		for col in [3,7,8,9,10,11,12,16,17,18,20,21,22,23,24,25,26]:
			for row in range(6, len(self.Item_List) + 6):  
				cell = ws.cell(row=row, column=col)
				cell.number_format = '#,##0.00'

		for col in [13,15]:
			for row in range(6, len(self.Item_List)+6):  
				cell = ws.cell(row=row, column=col)
				cell.number_format = '#,##0'

		for col in range(1, ws.max_column + 1):
			for row in range(1, len(self.Item_List)+6):  
				cell = ws.cell(row=row, column=col)
				cell.font = Font(size=16)

		sign_row = len(self.Item_List)+7
		ws[f"B{sign_row}"] = "核准:"
		ws[f"H{sign_row}"] = "審核:"
		ws[f"M{sign_row}"] = "主管:"
		ws[f"Q{sign_row}"] = "稽核:"
		ws[f"V{sign_row}"] = "承辦:"

		for col in range(1, ws.max_column + 1): 
			cell = ws.cell(row=sign_row, column=col)
			cell.font = Font(size=16, bold=True)
			
		# Define the path for saving the new Excel file
		order_cost = os.path.join(self.PATH, "訂單成本表-更新.xlsx")

		# Save the workbook to the specified path
		wb.save(order_cost)


if __name__ == "__main__":
    bot = ORDER_COST_EXPORT()  # Instantiate the class
    bot.FILL_EXCEL_EXPORT()
    print("訂單成本表已更新至路徑")